# -*- coding: utf-8 -*-
"""
Background worker — runs permanently on Railway.
Picks up search jobs from Redis, runs them, saves results back to Redis.
Streamlit then sends the email (it has outbound network access).
"""
import json
import os
import time
import threading
import re
import io
import requests
import traceback
import redis
from base64 import b64encode
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import deque
from sic_data import SIC_LOOKUP

REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379")

def get_redis():
    return redis.from_url(
        REDIS_URL,
        decode_responses=True,
        socket_connect_timeout=5,
        socket_timeout=10,
        health_check_interval=30,
    )

def is_cancelled(job_id):
    """Check if a cancel has been requested for this job."""
    try:
        r = get_redis()
        return r.get("ch_cancel") == job_id
    except:
        return False

API_BASE = "https://api.company-information.service.gov.uk"

class RateLimiter:
    def __init__(self, max_calls=575, window=300):
        self.max_calls = max_calls
        self.window = window
        self.calls = deque()
        self.lock = threading.Lock()

    def record_call(self):
        with self.lock:
            self.calls.append(time.time())

    def calls_in_window(self):
        with self.lock:
            now = time.time()
            cutoff = now - self.window
            while self.calls and self.calls[0] < cutoff:
                self.calls.popleft()
            return len(self.calls)

    def wait_if_needed(self):
        while True:
            count = self.calls_in_window()
            if count < self.max_calls:
                self.record_call()
                return
            with self.lock:
                wait_time = (self.calls[0] + self.window) - time.time() + 0.1 if self.calls else 1
            time.sleep(max(0.1, wait_time))

_rl = RateLimiter()

def ch_get(path, api_key):
    _rl.wait_if_needed()
    auth = "Basic " + b64encode(f"{api_key}:".encode()).decode()
    # Tuple timeout: (connect_timeout, read_timeout)
    r = requests.get(API_BASE + path, headers={"Authorization": auth}, timeout=(5, 15))
    if r.status_code == 429:
        time.sleep(5)
        _rl.wait_if_needed()
        r = requests.get(API_BASE + path, headers={"Authorization": auth}, timeout=(5, 15))
    r.raise_for_status()
    return r.json()

def fmt_currency(val):
    if val is None: return ""
    try:
        v = float(val)
        if abs(v) >= 1_000_000: return f"£{v/1_000_000:.1f}m"
        elif abs(v) >= 1_000:   return f"£{v/1_000:.0f}k"
        else:                    return f"£{v:.0f}"
    except: return ""

def title_case_company(name):
    if not name: return name
    PRESERVE = {"UK","LLP","LTD","PLC","USA","IT","HR","PR"}
    words = name.split()
    result = []
    for i, w in enumerate(words):
        clean = w.strip(".,()&")
        if clean.upper() in PRESERVE and i > 0:
            result.append(clean.upper())
        else:
            result.append("-".join(p.capitalize() for p in w.split("-")))
    return " ".join(result)

def split_director_name(full_name):
    if not full_name: return "", ""
    TITLES = {"mr","mrs","ms","miss","dr","prof","sir","dame","rev","rt","hon",
              "lord","lady","cllr","capt","maj","col","lt","cmdr","qc","kc"}
    parts = full_name.strip().split()
    while parts and parts[0].lower().rstrip(".") in TITLES:
        parts = parts[1:]
    if not parts: return "", ""
    def tc(s): return "-".join(w.capitalize() for w in s.split("-"))
    caps  = [p for p in parts if p.replace("-","").isupper() and len(p) > 1]
    lower = [p for p in parts if not p.replace("-","").isupper()]
    if caps:
        surname = tc(caps[-1])
        first   = tc(lower[0]) if lower else tc(parts[0])
    else:
        first   = tc(parts[0])
        surname = tc(parts[-1]) if len(parts) > 1 else ""
    return first, surname

def fetch_financials(company_number, api_key):
    from bs4 import BeautifulSoup
    result = {"accounts_date":"","cash_at_bank":"","total_assets":"","net_assets":"",
              "fixed_assets":"","current_assets":"","employees":"","accountant":"",
              "business_address":"","is_dormant": False}
    try:
        auth = "Basic " + b64encode(f"{api_key}:".encode()).decode()
        headers = {"Authorization": auth}
        _rl.wait_if_needed()
        fh = requests.get(f"{API_BASE}/company/{company_number}/filing-history",
                         params={"category":"accounts","items_per_page":10},
                         headers=headers, timeout=(5, 8))
        if fh.status_code == 429:
            time.sleep(3); _rl.wait_if_needed()
            fh = requests.get(f"{API_BASE}/company/{company_number}/filing-history",
                             params={"category":"accounts","items_per_page":10},
                             headers=headers, timeout=(5, 8))
        if fh.status_code != 200: return result
        all_filings = fh.json().get("items",[])
        # If ALL filings are dormant, mark as dormant
        if all_filings and all("dormant" in f.get("description","").lower() for f in all_filings):
            result["is_dormant"] = True
            return result
        filings = [f for f in all_filings
                   if "dormant" not in f.get("description","").lower()]
        if not filings: return result
        latest = filings[0]
        result["accounts_date"] = latest.get("action_date", latest.get("date",""))
        doc_meta_url = latest.get("links",{}).get("document_metadata","")
        if not doc_meta_url: return result
        _rl.wait_if_needed()
        dm = requests.get(doc_meta_url, headers=headers, timeout=(5, 8))
        if dm.status_code != 200: return result
        meta = dm.json()
        doc_url = meta.get("links",{}).get("document","")
        if not doc_url: return result
        if "application/xhtml+xml" not in meta.get("resources",{}): return result
        # iXBRL doc fetch — wrapped in a daemon thread with a TRUE hard wall-clock timeout.
        # iter_content() blocks internally between chunks, so a wall-clock check inside
        # the loop never fires while waiting. A server drip-feeding data just under the
        # per-chunk read timeout can hold a thread indefinitely. thread.join(timeout=N)
        # always fires after N seconds regardless of what the thread is doing.
        _MAX_FETCH_BYTES = 15 * 1024 * 1024  # 15 MB size cap
        _HARD_TIMEOUT    = 30                  # seconds — absolute ceiling per document
        _content_holder  = [None]

        def _fetch_doc():
            try:
                _r = requests.get(doc_url,
                                  headers={**headers, "Accept": "application/xhtml+xml"},
                                  timeout=(5, 15), stream=True)
                if _r.status_code != 200:
                    _r.close()
                    return
                _chunks = []; _total = 0
                for _chunk in _r.iter_content(chunk_size=65536):
                    if _chunk:
                        _total += len(_chunk)
                        if _total > _MAX_FETCH_BYTES:
                            _r.close()
                            return
                        _chunks.append(_chunk)
                _r.close()
                _content_holder[0] = b"".join(_chunks)
            except Exception:
                pass

        _fetch_thread = threading.Thread(target=_fetch_doc, daemon=True)
        _fetch_thread.start()
        _fetch_thread.join(timeout=_HARD_TIMEOUT)
        if _content_holder[0] is None:
            return result
        _doc_content = _content_holder[0]
        _doc_text_preview = _doc_content[:2000].decode("utf-8", errors="ignore")
        # Check for dormant in the actual document content or URL
        if "dormant" in doc_url.lower() or "dormant" in _doc_text_preview.lower():
            result["is_dormant"] = True
            return result
        soup = BeautifulSoup(_doc_content, "html.parser")
        # Final check in parsed text (catches "accounts for a dormant company" headings)
        doc_text_sample = soup.get_text()[:3000].lower()
        if "dormant" in doc_text_sample:
            result["is_dormant"] = True
            return result

        def get_val(soup, tag_names):
            for tag_name in tag_names:
                for tag in soup.find_all(attrs={"name":True}):
                    name_attr = tag.get("name","")
                    bare = name_attr.split(":")[-1] if ":" in name_attr else name_attr
                    if bare.lower() != tag_name.lower(): continue
                    ctx = tag.get("contextref","")
                    if any(x in ctx.lower() for x in ["prior","previous","preceding"]): continue
                    sign = tag.get("sign","")
                    scale = int(tag.get("scale","0") or "0")
                    try:
                        raw = tag.get_text(strip=True).replace(",","").replace(" ","").replace("\xa0","")
                        if not raw or raw in ("-","—"): continue
                        val = float(raw) * (10**scale)
                        if sign == "-": val = -val
                        if val != 0: return val
                    except: continue
            return None

        def fv(v): return fmt_currency(v) if v is not None else ""
        def ev(v):
            if v is None: return ""
            try:
                i = int(round(float(v)))
                if i <= 0 or i > 5000 or (1980 <= i <= 2040): return ""
                return str(i)
            except: return ""

        # Employee special extraction — use raw integer ignoring scale
        def get_val_raw_int(soup, tag_names):
            """Like get_val but ignores scale — for employee counts."""
            for tag_name in tag_names:
                for tag in soup.find_all(attrs={"name":True}):
                    name_attr = tag.get("name","")
                    bare = name_attr.split(":")[-1] if ":" in name_attr else name_attr
                    if bare.lower() != tag_name.lower(): continue
                    ctx = tag.get("contextref","")
                    if any(x in ctx.lower() for x in ["prior","previous","preceding"]): continue
                    try:
                        raw = tag.get_text(strip=True).replace(",","").replace(" ","").replace(" ","")
                        if not raw or raw in ("-","—"): continue
                        val = int(float(raw))  # ignore scale completely
                        if val > 0: return val
                    except: continue
            return None

        result["total_assets"]   = fv(get_val(soup,["TotalAssetsLessCurrentLiabilities","TotalAssets","BalanceSheetTotal","Assets"]))
        result["net_assets"]     = fv(get_val(soup,["NetAssetsLiabilities","NetAssets","ShareholdersEquity","Equity","MembersOtherInterests","TotalMembersInterests","MembersCapital"]))
        result["fixed_assets"]   = fv(get_val(soup,["FixedAssets","TotalFixedAssets","NonCurrentAssets"]))
        result["current_assets"] = fv(get_val(soup,["CurrentAssets","TotalCurrentAssets"]))
        result["cash_at_bank"]    = fv(get_val(soup,["CashBankInHand","CashBankOnHand","Cash","CashAndCashEquivalents","CashAtBankAndInHand"]))
        emp_raw = get_val_raw_int(soup,["AverageNumberEmployeesDuringPeriod","NumberEmployees","AverageNumberPersonsEmployed","EmployeesTotal","NumberOfEmployees","AverageNumberOfEmployees","EmployeeCount","Staff","NumberStaff","AverageNumberStaff"])
        result["employees"] = str(emp_raw) if emp_raw and 0 < emp_raw < 5000 and not (1980 <= emp_raw <= 2040) else ""

        if not result["employees"]:
            text = soup.get_text().lower()
            MONTHS = ["january","february","march","april","may","june","july",
                      "august","september","october","november","december"]
            for pat in [r"average\s+number\s+of\s+(?:employees|persons\s+employed)\s+(?:during\s+the\s+(?:year|period)\s+)?(?:was|were|:)\s*(\d{1,4})",
                        r"number\s+of\s+employees[^.]{0,40}(?:was|were|:)\s*(\d{1,4})"]:
                m = re.search(pat, text)
                if m:
                    v = int(m.group(1))
                    surrounding = text[max(0,m.start(1)-30):m.start(1)+10]
                    if 0 < v < 1000 and not (1980 <= v <= 2040) and not any(mo in surrounding for mo in MONTHS):
                        result["employees"] = str(v)
                        break

        try:
            SUFFIXES = r"(?:LLP|Chartered Accountants|Certified Accountants|Chartered Certified Accountants|& Co(?:\.|mpany)?|Accountants)"
            trigger_pat = (r"(?:prepared by|statutory auditors?|reporting accountants?|"
                          r"independent auditors?|audited by|accounts? (?:have been )?prepared by)"
                          r"[:\s]+([A-Z][A-Za-z0-9 &,\.\-]{2,50}?" + SUFFIXES + r")")
            full_text = soup.get_text(separator=" ", strip=True)
            m = re.search(trigger_pat, full_text)
            if m:
                accountant = m.group(1).strip().rstrip(".,")
            else:
                fallback_pat = r"([A-Z][A-Za-z0-9 &,\.\-]{2,50}?" + SUFFIXES + r")"
                accountant = ""
                for m in re.finditer(fallback_pat, full_text):
                    candidate = m.group(1).strip().rstrip(".,")
                    skip = ["the company","the directors","companies house","hmrc",
                            "limited company","association of","institute of",
                            "liability partnership","recruitment","staffing",
                            "employment","personnel","limited liability",
                            "limitedliabilitypartnership","limited liability partnership llp"]
                    if any(s in candidate.lower() for s in skip): continue
                    if len(candidate) > 4:
                        accountant = candidate
                        break
            if accountant:
                accountant = accountant.split("|")[0].strip()
                for prefix in ["Pages For Filing With Registrar ","PAGES FOR FILING WITH REGISTRAR "]:
                    if accountant.startswith(prefix):
                        accountant = accountant[len(prefix):]
                accountant = accountant.strip()[:60]
            result["accountant"] = accountant
        except: pass

        # Business address — address tags excluding the registered office context (_7_8)
        try:
            def get_addr_parts(soup, tag_names):
                parts = []
                for tag_name in tag_names:
                    for tag in soup.find_all(attrs={"name": True}):
                        name_attr = tag.get("name", "")
                        bare = name_attr.split(":")[-1] if ":" in name_attr else name_attr
                        if bare.lower() != tag_name.lower(): continue
                        ctx = tag.get("contextref", "")
                        if "_7_8" in ctx or "registeredoffice" in ctx.lower(): continue
                        val = tag.get_text(strip=True)
                        if val: parts.append(val)
                        break
                return parts

            addr_parts = get_addr_parts(soup, [
                "AddressLine1", "AddressLine2", "AddressLine3",
                "PrincipalLocation-CityOrTown", "CountyRegion", "PostalCodeZip"
            ])
            if addr_parts:
                result["business_address"] = ", ".join(addr_parts)
        except: pass

    except: pass
    return result

def calc_score(fin):
    score = 0
    def parse_val(s):
        if s is None or s == "": return None
        # Handle both numeric (new) and string (legacy) formats
        if isinstance(s, (int, float)): return float(s)
        try:
            s = str(s).replace("£","").replace(",","").strip()
            neg = s.startswith("-"); s = s.lstrip("-")
            mult = 1_000_000 if s.endswith("m") else (1_000 if s.endswith("k") else 1)
            return float(s.rstrip("mk")) * mult * (-1 if neg else 1)
        except: return None
    na = parse_val(fin.get("net_assets",""))
    if na:
        if na > 500_000: score += 2
        elif na > 100_000: score += 1
    emp = fin.get("employees","")
    if emp:
        try:
            e = int(emp)
            if e >= 20: score += 2
            elif e >= 5: score += 1
        except: pass
    ca = parse_val(fin.get("current_assets",""))
    if ca and ca > 200_000: score += 1
    return score

def fetch_all_for_sic(sic_code, base_params, api_key):
    auth = "Basic " + b64encode(f"{api_key}:".encode()).decode()
    headers = {"Authorization": auth}
    params_base = {**base_params}
    if sic_code: params_base["sic_codes"] = sic_code
    items = []; start = 0; total = None
    while True:
        params = {**params_base, "size": 100, "start_index": start}
        _rl.wait_if_needed()
        r = requests.get(API_BASE + "/advanced-search/companies",
                        params=params, headers=headers, timeout=(5, 10))
        # Retry on 429 rate limit — up to 3 attempts with backoff
        if r.status_code == 429:
            for _attempt in range(3):
                time.sleep(5 * (_attempt + 1))
                _rl.wait_if_needed()
                r = requests.get(API_BASE + "/advanced-search/companies",
                                params=params, headers=headers, timeout=(5, 10))
                if r.status_code != 429: break
        if r.status_code not in (200,): break
        data = r.json()
        batch = data.get("items", data.get("companies",[]))
        if total is None: total = data.get("hits", data.get("total_results",0))
        if not batch: break
        items.extend(batch)
        start += len(batch)
        if start >= (total or 0) or start >= 5000: break
    return items

def write_status(status):
    try:
        r = get_redis()
        r.set("ch_status", json.dumps(status))
    except Exception as e:
        print(f"[{datetime.now()}] Status write error: {e}")

def run_job(job):
    api_key        = os.environ.get("CH_API_KEY","")
    sg_key         = os.environ.get("SENDGRID_API_KEY","")
    email_to       = job.get("email_to","")
    location       = job.get("location","Surrey")
    selected_sics  = job.get("sic_codes",[])
    sic_labels     = job.get("sic_labels",[])
    fetch_fin_flag = job.get("fetch_financials", True)
    min_age        = job.get("min_age", 3)
    max_age        = job.get("max_age", 0)
    excl_dormant   = job.get("excl_dormant", True)
    min_net_assets = job.get("min_net_assets", 0)
    emp_min        = job.get("emp_min", 0)
    emp_max        = job.get("emp_max", 0)
    one_per_co     = job.get("one_per_company", True)
    linkedin_hyperlinks = job.get("linkedin_hyperlinks", True)
    company_types  = job.get("company_types", ["ltd","llp"])

    # Clear any cancel flag from previous job
    try:
        get_redis().delete("ch_cancel")
    except: pass

    write_status({"running": True, "stage": "Fetching companies...", "dir_done": 0,
                  "fin_done": 0, "total": 0, "started_at": time.time(), "error": None,
                  "job_id": job.get("job_id",""), "ready_to_email": False})
    try:
        base_params = {"location": location, "company_status": "active"}
        if company_types: base_params["company_type"] = ",".join(company_types)

        # ── Event log — defined here so it's available throughout the entire job ──
        event_log = []

        def log_event(msg):
            """Append to event log and print to Railway deploy logs."""
            event_log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
            print(f"[{datetime.now()}] {msg}")

        def send_event_log_email(subject_suffix, extra_lines=None):
            """Send the event log as a plain-text email — called on timeout or error."""
            if not email_to or not sg_key: return
            try:
                import sendgrid as _sg
                from sendgrid.helpers.mail import Mail as _Mail
                lines = [
                    "Companies House Prospector — Event Log",
                    "=" * 52,
                    f"Location      : {location}",
                    f"SIC codes     : {len(selected_sics)}",
                    f"Net assets min: £{min_net_assets:,}" if min_net_assets else "Net assets min: None",
                    f"Companies     : {total:,}" if total else "Companies     : SIC search not yet complete",
                    "",
                    "EVENT LOG",
                    "─" * 40,
                ] + event_log
                if extra_lines:
                    lines += ["", "─" * 40] + extra_lines
                subject = f"Prospector — {subject_suffix} — {location}"
                msg = _Mail(from_email="sillarsdave@gmail.com", to_emails=email_to,
                            subject=subject, plain_text_content="\n".join(lines))
                _holder = [None]
                def _do():
                    try: _holder[0] = _sg.SendGridAPIClient(api_key=sg_key).send(msg)
                    except Exception as _e: print(f"[{datetime.now()}] Event log email error: {_e}")
                _t = threading.Thread(target=_do, daemon=True)
                _t.start(); _t.join(timeout=60)
                print(f"[{datetime.now()}] Event log email sent: {subject}")
            except Exception as _e:
                print(f"[{datetime.now()}] Failed to send event log email: {_e}")

        log_event(f"Job started — {location} | {len(selected_sics)} SIC codes"
                  + (f" | £{min_net_assets:,}+ net assets" if min_net_assets else "")
                  + (f" | {min_age}yr+ age" if min_age else ""))

        start_time = time.time()
        all_items = []; seen = set()
        for i, sic in enumerate(selected_sics):
            if is_cancelled(job.get("job_id","")):
                print(f"[{datetime.now()}] Job cancelled during SIC search")
                break
            fetched = fetch_all_for_sic(sic, base_params, api_key)
            for c in fetched:
                num = c.get("company_number","")
                if num and num not in seen:
                    seen.add(num); all_items.append(c)
            if (i + 1) % 5 == 0 or (i + 1) == len(selected_sics):
                write_status({"running": True,
                              "stage": f"Searching SIC codes ({i+1}/{len(selected_sics)} done)...",
                              "dir_done": 0, "fin_done": 0, "total": 0,
                              "started_at": start_time, "job_id": job.get("job_id",""),
                              "error": None, "ready_to_email": False})

        today = date.today()
        filtered = []
        for c in all_items:
            if excl_dormant and "dormant" in c.get("company_status","").lower(): continue
            inc = c.get("date_of_creation","")
            if inc:
                try:
                    y,m2,d2 = inc.split("-")
                    age_yrs = (today - date(int(y),int(m2),int(d2))).days // 365
                    if min_age > 0 and age_yrs < min_age: continue
                    if max_age > 0 and age_yrs > max_age: continue
                except: pass
            filtered.append(c)
        all_items = filtered
        total = len(all_items)

        log_event(f"SIC search complete — {total:,} companies (after age/dormant filter)")
        log_event(f"Starting pipeline: financials for all {total:,}, directors only for passing companies")

        write_status({"running": True, "stage": f"Loading directors and financials for {total:,} companies...",
                      "dir_done": 0, "fin_done": 0, "total": total,
                      "started_at": time.time(), "error": None, "ready_to_email": False})

        director_cache = {}; financials_cache = {}
        dir_lock = threading.Lock(); fin_lock = threading.Lock()
        dir_done = [0]; fin_done = [0]
        passing_companies = []          # companies that passed the filter
        start_time = time.time()
        _job_deadline = time.time() + 259200  # 72-hour overall ceiling

        # ── Shared helpers ────────────────────────────────────────────────────
        def _parse_fin_val(s):
            """Parse a formatted currency string (£1.2m, £500k, -£100) to float.
            Returns None if the value is missing or unparseable."""
            if s is None or s == "": return None
            try:
                _s = str(s).replace("£","").replace(",","").strip()
                _neg = _s.startswith("-"); _s = _s.lstrip("-")
                _mult = 1_000_000 if _s.endswith("m") else (1_000 if _s.endswith("k") else 1)
                return float(_s.rstrip("mk")) * _mult * (-1 if _neg else 1)
            except: return None

        def _passes_filter(c, fin):
            """Return True if this company passes all user filters.
            Called immediately when financials arrive — only passing companies
            get director fetches, saving ~90% of director API calls."""
            if excl_dormant and "dormant" in c.get("company_status","").lower(): return False
            if excl_dormant and fin.get("is_dormant", False): return False
            if min_net_assets > 0:
                na_val = _parse_fin_val(fin.get("net_assets", None))
                if na_val is not None:
                    if na_val < min_net_assets: return False
                else:
                    # No net assets — use cash at bank AND total assets as proxies
                    ca_val = _parse_fin_val(fin.get("cash_at_bank", None))
                    ta_val = _parse_fin_val(fin.get("total_assets", None))
                    if not (ca_val is not None and ca_val >= min_net_assets and
                            ta_val is not None and ta_val >= min_net_assets):
                        return False
            if emp_min > 0 or emp_max > 0:
                emp_s = fin.get("employees","")
                if emp_s:
                    try:
                        e = int(emp_s)
                        if emp_min > 0 and e < emp_min: return False
                        if emp_max > 0 and e > emp_max: return False
                    except: pass
            return True

        def _on_dir_complete(future):
            """Callback — runs immediately when a director future completes.
            Updates director_cache and dir_done from whichever worker thread
            finished the fetch, so dir_done reflects real progress during Phase 1."""
            try:
                num, active = future.result()
                with dir_lock:
                    director_cache[num] = active
            except Exception:
                pass
            with dir_lock:
                dir_done[0] += 1

        def fetch_dir(c):
            num = c.get("company_number","")
            if not num: return num, []
            try:
                d = ch_get(f"/company/{num}/officers?items_per_page=10", api_key)
                active = [o for o in d.get("items",[])
                          if not o.get("resigned_on") and
                          o.get("officer_role","") in ("director","llp-designated-member","member")]
                return num, active
            except: return num, []

        def fetch_fin(c):
            num = c.get("company_number","")
            if not num: return num, {}
            return num, fetch_financials(num, api_key)

        job_id_str = job.get("job_id","")
        fin_ex = ThreadPoolExecutor(max_workers=5)
        dir_ex = ThreadPoolExecutor(max_workers=6)
        dir_futures = {}   # future -> company (only for companies that passed filter)

        # ── Phase 1: Financials → filter → directors (all running concurrently) ─
        # fin_ex fetches financials for every company.
        # As each result arrives the filter is applied immediately.
        # Passing companies are submitted to dir_ex right away — so director
        # fetches run in parallel with the remaining financial fetches.
        # This means we only call the officers API for companies we'll actually use.
        fin_phase_start = time.time()
        try:
            if not fetch_fin_flag:
                # Financials disabled — all companies go straight to directors
                fin_done[0] = total
                for c in all_items:
                    passing_companies.append(c)
                    _df = dir_ex.submit(fetch_dir, c)
                    _df.add_done_callback(_on_dir_complete)
                    dir_futures[_df] = c
            else:
                fin_futures = {fin_ex.submit(fetch_fin, c): c for c in all_items}
                for fin_future in as_completed(fin_futures):
                    if is_cancelled(job_id_str):
                        print(f"[{datetime.now()}] Job cancelled during financials fetch")
                        break
                    if time.time() > _job_deadline or time.time() - fin_phase_start > 86400:
                        print(f"[{datetime.now()}] Financials timeout — proceeding with partial results")
                        break
                    c = fin_futures[fin_future]
                    try:
                        num, fin = fin_future.result()
                        with fin_lock: financials_cache[num] = fin
                        if _passes_filter(c, fin):
                            passing_companies.append(c)
                            _df = dir_ex.submit(fetch_dir, c)
                            _df.add_done_callback(_on_dir_complete)
                            dir_futures[_df] = c
                    except Exception:
                        pass
                    with fin_lock: fin_done[0] += 1
                    write_status({"running": True, "stage": "Loading directors and financials...",
                                  "dir_done": dir_done[0], "fin_done": fin_done[0],
                                  "total": total, "started_at": start_time,
                                  "job_id": job_id_str, "error": None, "ready_to_email": False})
                    if fin_done[0] % 500 == 0:
                        elapsed = (time.time() - start_time) / 3600
                        log_event(f"Financials: {fin_done[0]:,}/{total:,} "
                                  f"({fin_done[0]/total*100:.1f}%) | "
                                  f"{len(passing_companies):,} passed filter | "
                                  f"{elapsed:.1f}h elapsed")
        finally:
            try: fin_ex.shutdown(wait=False, cancel_futures=True)
            except Exception: pass

        if time.time() > _job_deadline or time.time() - fin_phase_start > 86400:
            log_event(f"TIMEOUT — Financials hit time limit at {fin_done[0]:,}/{total:,} "
                      f"| {len(passing_companies):,} passed filter")
            send_event_log_email("Financials Timeout — Partial Results",
                                 [f"Financials completed: {fin_done[0]:,}/{total:,}",
                                  f"Companies passed filter: {len(passing_companies):,}",
                                  "Job will continue building results from data collected so far."])
        else:
            log_event(f"Financials complete — {fin_done[0]:,}/{total:,} fetched "
                      f"| {len(passing_companies):,} passed filter")

        # ── Phase 2: Wait for any remaining director futures ─────────────────
        # Callbacks already collected results as they completed during Phase 1.
        # This loop just ensures all in-flight futures finish and writes status.
        log_event(f"Director collection: {len(dir_futures):,} passing companies to process")
        dir_phase_start = time.time()
        try:
            for dir_future in as_completed(dir_futures):
                if is_cancelled(job_id_str):
                    print(f"[{datetime.now()}] Job cancelled during directors fetch")
                    break
                if time.time() > _job_deadline or time.time() - dir_phase_start > 86400:
                    log_event(f"TIMEOUT — Directors hit time limit at {dir_done[0]:,}/{len(dir_futures):,}")
                    send_event_log_email("Directors Timeout — Partial Results",
                                         [f"Directors completed: {dir_done[0]:,}/{len(dir_futures):,}",
                                          "Job will continue building results from data collected so far."])
                    break
                # Data already handled by _on_dir_complete callback — just update status
                write_status({"running": True, "stage": "Loading directors and financials...",
                              "dir_done": dir_done[0], "fin_done": fin_done[0],
                              "total": total, "started_at": start_time,
                              "job_id": job_id_str, "error": None, "ready_to_email": False})
        finally:
            try: dir_ex.shutdown(wait=False, cancel_futures=True)
            except Exception: pass

        log_event(f"Directors complete — {dir_done[0]:,}/{len(dir_futures):,} fetched")

        if is_cancelled(job_id_str):
            print(f"[{datetime.now()}] Job cancelled — skipping results and email")
            write_status({"running": False, "stage": "Cancelled", "job_id": job_id_str,
                          "error": None, "email_sent": False, "ready_to_email": False})
            return

        write_status({"running": True, "stage": "Building results...",
                      "dir_done": dir_done[0], "fin_done": fin_done[0],
                      "total": total, "started_at": start_time,
                      "job_id": job_id_str,
                      "error": None, "ready_to_email": False})

        def sort_key(c):
            fin = financials_cache.get(c.get("company_number",""),{})
            score = calc_score(fin)
            na_val = _parse_fin_val(fin.get("net_assets", None))
            return (score, na_val if na_val is not None else -999999)

        # Companies already filtered by _passes_filter during the financial phase.
        # Just sort — no re-filtering needed.
        results = list(passing_companies)
        results.sort(key=sort_key, reverse=True)

        rows = []
        for c in results:
            num = c.get("company_number","")
            company_name = title_case_company(c.get("company_name", c.get("title","")))
            addr = c.get("registered_office_address",{})
            addr_str = " ".join(filter(None,[addr.get("address_line_1",""),
                                             addr.get("locality",""), addr.get("postal_code","")]))
            sics = "; ".join(c.get("sic_codes",[]))
            inc = c.get("date_of_creation","")
            age = ""
            if inc:
                try:
                    y,m2,d2 = inc.split("-")
                    age = (today-date(int(y),int(m2),int(d2))).days//365
                except: pass
            fin = financials_cache.get(num,{})
            score = calc_score(fin)
            score_str = "★" * min(score,5) if score > 0 else "☆"
            dirs = director_cache.get(num,[])
            if one_per_co and dirs: dirs = dirs[:1]
            rows_data = dirs if dirs else [None]
            category = ", ".join(
                SIC_LOOKUP.get(s, f"SIC {s}")
                for s in c.get("sic_codes", []) if s
            )
            for o in rows_data:
                name = appt = ""
                if o:
                    name = " ".join(reversed([p.strip() for p in o.get("name","").split(",")]))
                    appt = o.get("appointed_on","")
                first_n, last_n = split_director_name(name)
                ch_url = f"https://find-and-update.company-information.service.gov.uk/company/{num}"
                li_url = "https://www.linkedin.com/search/results/people/?keywords=" + requests.utils.quote(f"{first_n} {last_n} {linkedin_company_keyword(company_name)}")

                def _parse_numeric(s):
                    """Convert £1.0m / £500k to float for sorting."""
                    if not s: return None
                    try:
                        s = str(s).replace("£","").replace(",","").strip()
                        neg = s.startswith("-"); s = s.lstrip("-")
                        mult = 1_000_000 if s.endswith("m") else (1_000 if s.endswith("k") else 1)
                        return float(s.rstrip("mk")) * mult * (-1 if neg else 1)
                    except: return None

                def _parse_emp(s):
                    try: return int(str(s).strip()) if s else None
                    except: return None

                rows.append({
                    "Score": score_str, "First Name": first_n, "Surname": last_n,
                    "Company": company_name, "Type": {"ltd": "LTD", "llp": "LLP", "plc": "PLC", "private-limited-guarant-nsc": "LTD", "private-unlimited": "LTD"}.get(c.get("company_type","").lower(), c.get("company_type","").upper()),
                    "Category": category, "Incorporated": inc, "Age": age,
                    "Fixed Assets": _parse_numeric(fin.get("fixed_assets","")),
                    "Current Assets": _parse_numeric(fin.get("current_assets","")),
                    "Total Assets": _parse_numeric(fin.get("total_assets","")),
                    "Net Assets": _parse_numeric(fin.get("net_assets","")),
                    "Cash at Bank": _parse_numeric(fin.get("cash_at_bank","")),
                    "Employees": _parse_emp(fin.get("employees","")),
                    "Accounts Date": fin.get("accounts_date",""),
                    "Dir. Appointed": appt, "Accountant": fin.get("accountant",""),
                    "Registered Address": addr_str,
                    "Business Address": fin.get("business_address",""),
                    "CH Link": ch_url, "LinkedIn": li_url,
                })

        # Build Excel
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from collections import Counter
        import base64

        df = pd.DataFrame(rows)
        # Sort by net assets descending (numeric now so sorts correctly)
        if "Net Assets" in df.columns and len(df) > 0:
            df = df.sort_values("Net Assets", ascending=False, na_position="last").reset_index(drop=True)
        wb = Workbook(); ws = wb.active; ws.title = "Prospects"
        _addr_cols = [c for c in ["Registered Address", "Business Address"] if c in df.columns]
        base_cols = [c for c in df.columns if c not in ["CH Link","LinkedIn"] + _addr_cols]
        headers_xl = base_cols + ["CH company","Officers","LinkedIn"] + _addr_cols
        CURRENCY_COLS = {"Total Assets","Net Assets","Fixed Assets","Current Assets","Cash at Bank"}
        NUMBER_COLS = {"Employees","Age"}
        hdr_fill = PatternFill("solid", fgColor="1a4a2e")
        for i, h in enumerate(headers_xl, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = Font(name="Arial", color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        fill_even = PatternFill("solid", fgColor="EBF3FB")
        fill_odd  = PatternFill("solid", fgColor="FFFFFF")
        for rn, (_, row) in enumerate(df.iterrows(), 2):
            fill = fill_even if rn % 2 == 0 else fill_odd
            row_vals = [row[c] for c in base_cols] + [row["CH Link"], f"{row['CH Link']}/officers", row["LinkedIn"]] + [row[c] for c in _addr_cols]
            for ci, val in enumerate(row_vals, 1):
                h = headers_xl[ci-1] if ci <= len(headers_xl) else ""
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.fill = fill; cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if h in CURRENCY_COLS and val is not None:
                    cell.number_format = '£#,##0;-£#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif h in NUMBER_COLS and val is not None:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                _link_start = len(base_cols) + 1
                _link_end   = len(base_cols) + 3
                if _link_start <= ci <= _link_end:
                    link_index = ci - len(base_cols) - 1
                    labels = ["Open", "Officers", "LinkedIn"]
                    urls = [row["CH Link"], f"{row['CH Link']}/officers", row["LinkedIn"]]
                    if link_index == 2 and linkedin_hyperlinks:
                        # LinkedIn as clickable hyperlink
                        cell.value = labels[link_index]
                        cell.hyperlink = urls[link_index]
                        cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                    else:
                        # CH company, Officers, and optionally LinkedIn — plain URL
                        cell.value = urls[link_index]
                        cell.font = Font(name="Arial", size=9)
        for ci, h in enumerate(headers_xl, 1):
            col_letter = get_column_letter(ci)
            max_len = len(str(h))
            for rn in range(2, ws.max_row+1):
                v = ws.cell(row=rn, column=ci).value
                if v: max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(max_len+2, 8), 40)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        ws_acct = wb.create_sheet("Accountants")
        acct_counts = Counter(r for r in df["Accountant"].tolist()
                              if r and str(r).strip()
                              and "audit" not in str(r).lower()
                              and str(r).strip().lower() not in {"n/a", "none", "not applicable"}
                              and len(str(r).strip()) > 3)
        for ci, h in enumerate(["Accountant Firm","No. of Clients","Companies"], 1):
            cell = ws_acct.cell(row=1, column=ci, value=h)
            cell.fill = PatternFill("solid", fgColor="1a4a2e")
            cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        acct_companies = {}
        for _, row in df.iterrows():
            acct = str(row.get("Accountant","")).strip()
            if acct:
                acct_companies.setdefault(acct, [])
                co = str(row.get("Company","")).strip()
                if co and co not in acct_companies[acct]: acct_companies[acct].append(co)
        for rn, (acct, count) in enumerate(acct_counts.most_common(), 2):
            ws_acct.cell(row=rn, column=1, value=acct).font = Font(name="Arial", size=9)
            ws_acct.cell(row=rn, column=2, value=count).font = Font(name="Arial", size=9)
            ws_acct.cell(row=rn, column=3, value=", ".join(acct_companies.get(acct,[]))).font = Font(name="Arial", size=9)
        ws_acct.column_dimensions["A"].width = 40
        ws_acct.column_dimensions["B"].width = 14
        ws_acct.column_dimensions["C"].width = 60

        end_time = time.time()
        _duration_secs = int(end_time - start_time)
        _duration_str = f"{_duration_secs // 3600}h {(_duration_secs % 3600) // 60}m {_duration_secs % 60}s"
        _start_dt = datetime.fromtimestamp(start_time)
        _end_dt = datetime.fromtimestamp(end_time)

        ws2 = wb.create_sheet("Search Criteria")
        _company_types_str = ", ".join([t.upper() for t in company_types]) if company_types else "All"
        _age_str = f"{min_age}yr+" if min_age and not max_age else (f"{min_age}–{max_age}yrs" if min_age and max_age else "Any")
        _emp_str = f"{emp_min}–{emp_max}" if (emp_min or emp_max) else "Any"
        criteria = {
            "Location": location,
            "Industries": ", ".join(sic_labels),
            "Company types": _company_types_str,
            "Min age": _age_str,
            "Exclude dormant": "Yes" if excl_dormant else "No",
            "Min net assets": f"£{min_net_assets:,}" if min_net_assets else "None",
            "Employees": _emp_str,
            "Fetch financials": "Yes" if fetch_fin_flag else "No",
            "One contact per company": "Yes" if one_per_co else "No",
            "Companies found": f"{total:,}",
            "After filters": f"{len(results):,}",
            "Results in export": f"{len(rows):,}",
            "Export date": today.strftime("%d %B %Y"),
            "Search started": _start_dt.strftime("%d %B %Y %H:%M:%S"),
            "Search ended": _end_dt.strftime("%d %B %Y %H:%M:%S"),
            "Duration": _duration_str,
        }
        for i, (k,v) in enumerate(criteria.items(), 1):
            ws2.cell(row=i, column=1, value=k).font = Font(bold=True, name="Arial")
            ws2.cell(row=i, column=2, value=str(v)).font = Font(name="Arial")

        xl_buf = io.BytesIO(); wb.save(xl_buf); xl_buf.seek(0)

        # Build CSV
        csv_df = df.copy()
        csv_df["CH company"] = df["CH Link"]
        csv_df["Officers"] = df["CH Link"].apply(lambda x: x+"/officers")
        csv_df["LinkedIn search"] = df["LinkedIn"]
        csv_df = csv_df.drop(columns=["CH Link","LinkedIn"])
        csv_str = csv_df.to_csv(index=False)

        # Save results to Redis FIRST (download fallback, 7-day expiry)
        search_date = today.strftime("%d %B %Y")
        loc_str = location.strip().replace(" ","_").lower()[:15]
        search_num = job.get("search_number", "")
        try:
            _r = get_redis()
            _r.set("ch_results_excel", base64.b64encode(xl_buf.getvalue()).decode(), ex=604800)
            _r.set("ch_results_csv", csv_str, ex=604800)
            _r.set("ch_results_meta", json.dumps({
                "search_date": search_date,
                "results_count": len(rows),
                "location": location,
                "industries": ", ".join(sic_labels),
                "job_id": job.get("job_id",""),
            }), ex=604800)
            print(f"[{datetime.now()}] Results saved to Redis for download")
        except Exception as _re:
            print(f"[{datetime.now()}] Redis save error: {_re}")

        # Send email via SendGrid (HTTPS - works on Railway)
        criteria = {"Location": location, "Industries": ", ".join(sic_labels),
                    "Total results": len(rows), "Export date": search_date}

        sg_key = os.environ.get("SENDGRID_API_KEY", "")
        from_email = "sillarsdave@gmail.com"
        SIZE_LIMIT = 20 * 1024 * 1024  # 20MB

        def build_excel(row_subset, part_label=""):
            """Build Excel workbook for a subset of rows, return bytes."""
            from openpyxl import Workbook as WB2
            from openpyxl.styles import Font as F2, PatternFill as PF2, Alignment as AL2
            from openpyxl.utils import get_column_letter as gcl2
            from collections import Counter as C2
            sub_df = pd.DataFrame(row_subset)
            wb2 = WB2(); ws2 = wb2.active
            ws2.title = "Prospects"
            _addr_cols2 = [c for c in ["Registered Address", "Business Address"] if c in sub_df.columns]
            base_cols2 = [c for c in sub_df.columns if c not in ["CH Link","LinkedIn"] + _addr_cols2]
            headers2 = base_cols2 + ["CH company","Officers","LinkedIn"] + _addr_cols2
            hf2 = PF2("solid", fgColor="1a4a2e")
            for i, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=i, value=h)
                cell.fill = hf2
                cell.font = F2(name="Arial", color="FFFFFF", bold=True, size=10)
                cell.alignment = AL2(horizontal="center", wrap_text=True)
            ws2.row_dimensions[1].height = 30
            fe2 = PF2("solid", fgColor="EBF3FB"); fo2 = PF2("solid", fgColor="FFFFFF")
            for rn, (_, row) in enumerate(sub_df.iterrows(), 2):
                fill = fe2 if rn%2==0 else fo2
                rv = [row[c] for c in base_cols2] + [row["CH Link"], f"{row['CH Link']}/officers", row["LinkedIn"]] + [row[c] for c in _addr_cols2]
                for ci, val in enumerate(rv, 1):
                    cell = ws2.cell(row=rn, column=ci, value=val)
                    cell.fill = fill; cell.font = F2(name="Arial", size=9)
                    cell.alignment = AL2(horizontal="left", vertical="center")
                    _ls2 = len(base_cols2) + 1
                    _le2 = len(base_cols2) + 3
                    if _ls2 <= ci <= _le2:
                        link_index = ci - len(base_cols2) - 1
                        lbls = ["Open", "Officers", "LinkedIn"]
                        urls2 = [row["CH Link"], f"{row['CH Link']}/officers", row["LinkedIn"]]
                        if link_index == 2 and linkedin_hyperlinks:
                            cell.value = lbls[link_index]
                            cell.hyperlink = urls2[link_index]
                            cell.font = F2(name="Arial", size=9, color="0563C1", underline="single")
                        else:
                            cell.value = urls2[link_index]
                            cell.font = F2(name="Arial", size=9)
            for ci, h in enumerate(headers2, 1):
                col_letter = gcl2(ci)
                max_len = max(len(str(h)), max((len(str(ws2.cell(row=rn2,column=ci).value or "")) for rn2 in range(2,ws2.max_row+1)), default=0))
                ws2.column_dimensions[col_letter].width = min(max(max_len+2,8),40)
            ws2.auto_filter.ref = ws2.dimensions
            ws2.freeze_panes = "A2"
            # Criteria sheet
            wsc = wb2.create_sheet("Search Criteria")
            crit2 = {**criteria, "Part": part_label} if part_label else criteria
            for i, (k,v) in enumerate(crit2.items(), 1):
                wsc.cell(row=i, column=1, value=k).font = F2(bold=True, name="Arial")
                wsc.cell(row=i, column=2, value=str(v)).font = F2(name="Arial")
            buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
            return buf2.getvalue()

        def send_sg(to_email, subject, body_text, xl_bytes, csv_bytes, date_str, part=""):
            suffix = f"_part{part}" if part else ""
            _sn = f"_search_{search_num}" if search_num else ""
            _fname_base = f"prospector_results_{loc_str}_{date_str.replace(' ','_')}{_sn}{suffix}"
            import sendgrid as sg_module
            from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
            msg = Mail(from_email=from_email, to_emails=to_email,
                       subject=subject, plain_text_content=body_text)
            msg.attachment = Attachment(FileContent(base64.b64encode(xl_bytes).decode()),
                FileName(f"{_fname_base}.xlsx"),
                FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                Disposition("attachment"))
            if csv_bytes:
                msg.attachment = Attachment(FileContent(base64.b64encode(csv_bytes).decode()),
                    FileName(f"{_fname_base}.csv"),
                    FileType("text/csv"), Disposition("attachment"))
            sg_client = sg_module.SendGridAPIClient(api_key=sg_key)
            # Wrap send in a thread to prevent hangs (SendGrid lib has no built-in timeout)
            _send_result = {"resp": None, "err": None}
            def _do_send():
                try: _send_result["resp"] = sg_client.send(msg)
                except Exception as _e: _send_result["err"] = _e
            _send_thread = threading.Thread(target=_do_send, daemon=True)
            _send_thread.start()
            _send_thread.join(timeout=900)  # 15-minute cap on email send
            if _send_thread.is_alive():
                print(f"[{datetime.now()}] WARNING: SendGrid send timed out after 15 minutes")
                return False
            if _send_result["err"]:
                print(f"[{datetime.now()}] SendGrid error: {_send_result['err']}")
                return False
            resp = _send_result["resp"]
            return resp.status_code in (200, 202)

        try:
            xl_bytes = xl_buf.getvalue()
            csv_bytes = csv_str.encode("utf-8-sig")
            total_size = len(xl_bytes) + len(csv_bytes)

            log_event(f"Results built — {len(results):,} companies | {len(rows):,} export rows")
            log_event(f"Building email — file size: {total_size/1024/1024:.1f}MB")

            body_base = ["Your Companies House Prospector search has completed.", ""]
            for k, v in criteria.items():
                body_base.append(f"{k}: {v}")
            body_base += ["", "─" * 40, "EVENT LOG", "─" * 40] + event_log + [""]

            if total_size <= SIZE_LIMIT:
                # Single email
                body = "\n".join(body_base + ["Please find the Excel and CSV results attached."])
                subject = f"Companies House Prospector Results — {search_date}"
                email_sent = send_sg(email_to, subject, body, xl_bytes, csv_bytes, search_date)
                print(f"[{datetime.now()}] Single email sent: {email_sent}")
            else:
                # Split into two parts
                print(f"[{datetime.now()}] File size {total_size/1024/1024:.1f}MB > 20MB — splitting into 2 emails")
                mid = len(rows) // 2
                parts = [rows[:mid], rows[mid:]]
                email_sent = True
                for i, part_rows in enumerate(parts, 1):
                    part_label = f"Part {i} of 2"
                    part_xl = build_excel(part_rows, part_label)
                    part_df = pd.DataFrame(part_rows)
                    part_df2 = part_df.copy()
                    part_df2["CH company"] = part_df["CH Link"]
                    part_df2["Officers"] = part_df["CH Link"].apply(lambda x: x+"/officers")
                    part_df2["LinkedIn search"] = part_df["LinkedIn"]
                    part_df2 = part_df2.drop(columns=["CH Link","LinkedIn"])
                    part_csv = part_df2.to_csv(index=False).encode("utf-8-sig")
                    body = "\n".join(body_base + [f"This is {part_label} ({len(part_rows):,} results). Please find attached."])
                    subject = f"Companies House Prospector Results — {search_date} ({part_label})"
                    ok = send_sg(email_to, subject, body, part_xl, part_csv, search_date, str(i))
                    if not ok:
                        email_sent = False
                    print(f"[{datetime.now()}] Part {i} email sent: {ok}")

        except Exception as email_err:
            email_sent = False
            print(f"[{datetime.now()}] Email error: {email_err}")

        write_status({"running": False, "stage": "Complete",
                      "job_id": job.get("job_id",""),
                      "dir_done": dir_done[0], "fin_done": fin_done[0],
                      "total": total, "started_at": start_time,
                      "completed_at": time.time(), "results_count": len(rows),
                      "ready_to_email": False, "email_sent": email_sent, "error": None})

        log_event(f"Job complete — {len(rows):,} results | email_sent={email_sent}")
        print(f"[{datetime.now()}] Job complete — {len(rows)} results, email_sent={email_sent}")

    except Exception as e:
        log_event(f"ERROR: {str(e)}")
        write_status({"running": False, "stage": "Error", "error": str(e),
                      "traceback": traceback.format_exc(), "ready_to_email": False})
        print(f"[{datetime.now()}] Job error: {e}")
        send_event_log_email("Job Error",
                             [f"Error: {str(e)}", "", "Traceback:", traceback.format_exc()])



def linkedin_company_keyword(company_name):
    """Return first 1-2 meaningful words of company name for LinkedIn search.
    Uses 2 words unless the second word is a legal suffix, in which case uses 1."""
    if not company_name:
        return ""
    SUFFIXES = {"limited","ltd","llp","plc","and co","company","group",
                "holdings","holding","services","solutions","consulting","consultancy",
                "management","associates","partnership","enterprises","ventures",
                "international","global","uk","the"}
    orig_words = company_name.split()
    if len(orig_words) >= 2 and orig_words[1].lower().rstrip(".") in SUFFIXES:
        return orig_words[0]
    return " ".join(orig_words[:min(2, len(orig_words))])

# ── Main loop ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"[{datetime.now()}] Worker started — connecting to Redis...")
    try:
        r = get_redis()
        r.ping()
        print(f"[{datetime.now()}] Redis connected OK")
    except Exception as e:
        print(f"[{datetime.now()}] Redis connection failed: {e}")

    last_job_id = None
    while True:
        try:
            r = get_redis()
            data = r.get("ch_job")
            if data:
                job = json.loads(data)
                job_id = job.get("job_id")
                if job_id and job_id != last_job_id:
                    # Check if this job already completed successfully
                    try:
                        status_raw = r.get("ch_status")
                        if status_raw:
                            status = json.loads(status_raw)
                            if status.get("job_id") == job_id and status.get("email_sent"):
                                print(f"[{datetime.now()}] Skipping already-completed job {job_id}")
                                last_job_id = job_id
                                continue
                    except: pass

                    last_job_id = job_id
                    print(f"[{datetime.now()}] New job: {job_id} — {job.get('location')} | {len(job.get('sic_codes',[]))} SIC codes")
                    run_job(job)
                    print(f"[{datetime.now()}] Job {job_id} complete")
        except Exception as e:
            print(f"[{datetime.now()}] Worker loop error: {e}")
        time.sleep(3)
