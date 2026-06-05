# -*- coding: utf-8 -*-
import streamlit as st
import requests
import time
import threading
import io
import re
import redis
import json
import base64
from base64 import b64encode
from datetime import date
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import deque
import os
import traceback
from datetime import datetime

_log_buffer = []

def get_log_text():
    return "\n".join(_log_buffer)

def get_redis():
    return redis.from_url(os.environ.get("REDIS_URL", "redis://localhost:6379"), decode_responses=True)

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Companies House Prospector",
    page_icon="\U0001f3e2",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .stApp { background: #f7f6f2; }
    .main-header {
        background: #1a4a2e;
        color: white;
        padding: 1rem 1.5rem;
        border-radius: 8px;
        margin-bottom: 1rem;
        font-family: Georgia, serif;
        font-size: 1.4rem;
        font-weight: bold;
    }
    div[data-testid="stSidebar"] { background: white; }
    .stButton > button {
        background: #1a4a2e;
        color: white;
        border: none;
        border-radius: 6px;
        font-family: Georgia, serif;
    }
    .stButton > button:hover { background: #2d6b47; }
</style>
""", unsafe_allow_html=True)

API_BASE = "https://api.company-information.service.gov.uk"

SIC_OPTIONS = [
    ("Accounting / bookkeeping (69201)", "69201"),
    ("Advertising agencies (73110)", "73110"),
    ("Antiques (47791)", "47791"),
    ("Architectural activities (71111)", "71111"),
    ("Architecture (71112)", "71112"),
    ("Art galleries retail (47781)", "47781"),
    ("Auditing (69202)", "69202"),
    ("Barristers (69101)", "69101"),
    ("Building construction (41202)", "41202"),
    ("Car dealerships (45111)", "45111"),
    ("Care agencies / staffing (78300)", "78300"),
    ("Catering / event catering (56210)", "56210"),
    ("Cosmetic / plastic surgery clinics (86102)", "86102"),
    ("Craft brewing (11050)", "11050"),
    ("Dental practices (86220)", "86220"),
    ("Distilling / spirits (11010)", "11010"),
    ("Engineering consultancy (71121)", "71121"),
    ("Estate agents (68310)", "68310"),
    ("Event management - conferences (82302)", "82302"),
    ("Event management - exhibitions (82301)", "82301"),
    ("General building contractors (41201)", "41201"),
    ("GP practices (86210)", "86210"),
    ("HR consulting (70229)", "70229"),
    ("Immigration consultants (69109)", "69109"),
    ("Insurance brokers (66220)", "66220"),
    ("IT consultancy (62020)", "62020"),
    ("Management consulting (70229)", "70229"),
    ("Market research (73200)", "73200"),
    ("Mortgage brokers (66190)", "66190"),
    ("Opticians (86230)", "86230"),
    ("Physical wellbeing / spa (96040)", "96040"),
    ("Physiotherapy (86901)", "86901"),
    ("PR / communications (70210)", "70210"),
    ("Property buying & selling (68100)", "68100"),
    ("Property development (41100)", "41100"),
    ("Property management (68320)", "68320"),
    ("Quantity surveying (74902)", "74902"),
    ("Racehorse owners (93191)", "93191"),
    ("Recruitment (78200)", "78200"),
    ("Software development (62012)", "62012"),
    ("Solicitors (69102)", "69102"),
    ("Sound recording / music publishing (59200)", "59200"),
    ("Surveying (71112)", "71112"),
    ("Tax consulting (69203)", "69203"),
    ("Tour operators (79120)", "79120"),
    ("Travel agencies (79110)", "79110"),
    ("TV programme production (59113)", "59113"),
    ("Veterinary (75000)", "75000"),
    ("Video production (59112)", "59112"),
    ("Vocational training (85320)", "85320"),
    ("Management consultancy (70210)", "70210"),
    ("Business support services (82190)", "82190"),
    ("Translation / interpretation (74300)", "74300"),
    ("Electrical installation (43210)", "43210"),
    ("Plumbing / heating (43220)", "43220"),
    ("Carpentry / joinery (43320)", "43320"),
    ("Painting / decorating (43341)", "43341"),
    ("Landscaping (81300)", "81300"),
    ("Specialist construction (43990)", "43990"),
    ("Osteopathy / chiropractic (86901)", "86901"),
    ("Graphic design (74101)", "74101"),
    ("IT support / infrastructure (62090)", "62090"),
]

class RateLimiter:
    def __init__(self, max_calls=575, window=300):
        self.max_calls = max_calls
        self.window = window
        self.calls = deque()
        self.lock = threading.Lock()

    def record_call(self):
        with self.lock:
            self.calls.append(time.time())

    def calls_in_window(self):
        with self.lock:
            now = time.time()
            cutoff = now - self.window
            while self.calls and self.calls[0] < cutoff:
                self.calls.popleft()
            return len(self.calls)

    def wait_if_needed(self):
        while True:
            count = self.calls_in_window()
            if count < self.max_calls:
                self.record_call()
                return
            with self.lock:
                wait_time = (self.calls[0] + self.window) - time.time() + 0.1 if self.calls else 1
            time.sleep(max(0.1, wait_time))

_rate_limiter = RateLimiter()

def ch_get(path, api_key):
    _rate_limiter.wait_if_needed()
    auth = "Basic " + b64encode(f"{api_key}:".encode()).decode()
    r = requests.get(API_BASE + path, headers={"Authorization": auth}, timeout=20)
    if r.status_code == 429:
        time.sleep(5)
        _rate_limiter.wait_if_needed()
        r = requests.get(API_BASE + path, headers={"Authorization": auth}, timeout=20)
    r.raise_for_status()
    return r.json()

def fmt_currency(val):
    if val is None: return ""
    try:
        v = float(val)
        if abs(v) >= 1_000_000: return f"\u00a3{v/1_000_000:.1f}m"
        elif abs(v) >= 1_000:   return f"\u00a3{v/1_000:.0f}k"
        else:                    return f"\u00a3{v:.0f}"
    except: return ""

def title_case_company(name):
    if not name: return name
    PRESERVE = {"UK","LLP","LTD","PLC","USA","IT","HR","PR"}
    words = name.split()
    result = []
    for i, w in enumerate(words):
        clean = w.strip(".,()&")
        if clean.upper() in PRESERVE and i > 0:
            result.append(clean.upper())
        else:
            result.append("-".join(p.capitalize() for p in w.split("-")))
    return " ".join(result)

def split_director_name(full_name):
    if not full_name: return "", ""
    TITLES = {"mr","mrs","ms","miss","dr","prof","sir","dame","rev","rt","hon",
              "lord","lady","cllr","capt","maj","col","lt","cmdr","qc","kc"}
    parts = full_name.strip().split()
    while parts and parts[0].lower().rstrip(".") in TITLES:
        parts = parts[1:]
    if not parts: return "", ""
    def tc(s): return "-".join(w.capitalize() for w in s.split("-"))
    caps = [p for p in parts if p.replace("-","").isupper() and len(p) > 1]
    lower = [p for p in parts if not p.replace("-","").isupper()]
    if caps:
        surname = tc(caps[-1])
        first = tc(lower[0]) if lower else tc(parts[0])
    else:
        first = tc(parts[0])
        surname = tc(parts[-1]) if len(parts) > 1 else ""
    return first, surname

def fetch_financials(company_number, api_key):
    result = {"accounts_date":"","turnover":"","total_assets":"","net_assets":"",
              "fixed_assets":"","current_assets":"","employees":"","accountant":""}
    try:
        from bs4 import BeautifulSoup
        auth = "Basic " + b64encode(f"{api_key}:".encode()).decode()
        headers = {"Authorization": auth}
        _rate_limiter.wait_if_needed()
        fh = requests.get(f"{API_BASE}/company/{company_number}/filing-history",
                         params={"category":"accounts","items_per_page":10},
                         headers=headers, timeout=12)
        if fh.status_code == 429:
            time.sleep(3); _rate_limiter.wait_if_needed()
            fh = requests.get(f"{API_BASE}/company/{company_number}/filing-history",
                             params={"category":"accounts","items_per_page":10},
                             headers=headers, timeout=12)
        if fh.status_code != 200: return result
        filings = [f for f in fh.json().get("items",[])
                   if "dormant" not in f.get("description","").lower()]
        if not filings: return result
        latest = filings[0]
        result["accounts_date"] = latest.get("action_date", latest.get("date",""))
        doc_meta_url = latest.get("links",{}).get("document_metadata","")
        if not doc_meta_url: return result
        _rate_limiter.wait_if_needed()
        dm = requests.get(doc_meta_url, headers=headers, timeout=12)
        if dm.status_code != 200: return result
        meta = dm.json()
        doc_url = meta.get("links",{}).get("document","")
        if not doc_url: return result
        if "application/xhtml+xml" not in meta.get("resources",{}): return result
        doc_r = requests.get(doc_url, headers={**headers,"Accept":"application/xhtml+xml"}, timeout=25)
        if doc_r.status_code != 200: return result
        soup = BeautifulSoup(doc_r.content, "html.parser")

        def get_val(soup, tag_names):
            for tag_name in tag_names:
                for tag in soup.find_all(attrs={"name":True}):
                    name_attr = tag.get("name","")
                    bare = name_attr.split(":")[-1] if ":" in name_attr else name_attr
                    if bare.lower() != tag_name.lower(): continue
                    ctx = tag.get("contextref","")
                    if any(x in ctx.lower() for x in ["prior","previous","preceding"]): continue
                    sign = tag.get("sign","")
                    scale = int(tag.get("scale","0") or "0")
                    try:
                        raw = tag.get_text(strip=True).replace(",","").replace(" ","").replace("\xa0","")
                        if not raw or raw in ("-","—"): continue
                        val = float(raw) * (10**scale)
                        if sign == "-": val = -val
                        if val != 0: return val
                    except: continue
            return None

        def fv(v): return fmt_currency(v) if v is not None else ""
        def ev(v):
            if v is None: return ""
            try:
                i = int(float(v))
                if i <= 0 or i > 5000 or (1980 <= i <= 2040): return ""
                return str(i)
            except: return ""

        result["total_assets"]   = fv(get_val(soup,["TotalAssetsLessCurrentLiabilities","TotalAssets","BalanceSheetTotal","Assets"]))
        result["net_assets"]     = fv(get_val(soup,["NetAssetsLiabilities","NetAssets","ShareholdersEquity","Equity"]))
        result["fixed_assets"]   = fv(get_val(soup,["FixedAssets","TotalFixedAssets","NonCurrentAssets"]))
        result["current_assets"] = fv(get_val(soup,["CurrentAssets","TotalCurrentAssets"]))
        result["employees"]      = ev(get_val(soup,["AverageNumberEmployeesDuringPeriod","NumberEmployees","AverageNumberPersonsEmployed"]))

        if not result["employees"]:
            text = soup.get_text().lower()
            MONTHS = ["january","february","march","april","may","june","july",
                      "august","september","october","november","december"]
            for pat in [r"average\s+number\s+of\s+(?:employees|persons\s+employed)\s+(?:during\s+the\s+(?:year|period)\s+)?(?:was|were|:)\s*(\d{1,4})",
                        r"number\s+of\s+employees[^.]{0,40}(?:was|were|:)\s*(\d{1,4})"]:
                m = re.search(pat, text)
                if m:
                    v = int(m.group(1))
                    surrounding = text[max(0,m.start(1)-30):m.start(1)+10]
                    if 0 < v < 1000 and not (1980 <= v <= 2040) and not any(mo in surrounding for mo in MONTHS):
                        result["employees"] = str(v)
                        break

        try:
            import re as _re
            full_text = soup.get_text(separator=" ", strip=True)
            accountant = ""
            SUFFIXES = r"(?:LLP|Chartered Accountants|Certified Accountants|Chartered Certified Accountants|& Co(?:\.|mpany)?|Accountants)"
            trigger_pat = (r"(?:prepared by|statutory auditors?|reporting accountants?|"
                          r"independent auditors?|audited by|accounts? (?:have been )?prepared by)"
                          r"[:\s]+([A-Z][A-Za-z0-9 &,\.\-]{2,50}?" + SUFFIXES + r")")
            m = _re.search(trigger_pat, full_text)
            if m:
                accountant = m.group(1).strip().rstrip(".,")
            else:
                fallback_pat = r"([A-Z][A-Za-z0-9 &,\.\-]{2,50}?" + SUFFIXES + r")"
                for m in _re.finditer(fallback_pat, full_text):
                    candidate = m.group(1).strip().rstrip(".,")
                    skip = ["the company","the directors","companies house","hmrc",
                            "limited company","association of","institute of",
                            "liability partnership","recruitment","staffing",
                            "employment","personnel","limited liability"]
                    if any(s in candidate.lower() for s in skip): continue
                    if len(candidate) > 4:
                        accountant = candidate
                        break
            if accountant:
                accountant = accountant.split("|")[0].strip()
                for prefix in ["Pages For Filing With Registrar ","PAGES FOR FILING WITH REGISTRAR "]:
                    if accountant.startswith(prefix):
                        accountant = accountant[len(prefix):]
                accountant = accountant.strip()[:60]
            result["accountant"] = accountant
        except: pass
    except: pass
    return result

def calc_score(fin):
    score = 0
    def parse_val(s):
        if not s: return None
        try:
            s = s.replace("\u00a3","").replace(",","").strip()
            neg = s.startswith("-"); s = s.lstrip("-")
            mult = 1_000_000 if s.endswith("m") else (1_000 if s.endswith("k") else 1)
            return float(s.rstrip("mk")) * mult * (-1 if neg else 1)
        except: return None
    na = parse_val(fin.get("net_assets",""))
    if na:
        if na > 500_000: score += 2
        elif na > 100_000: score += 1
    emp = fin.get("employees","")
    if emp:
        try:
            e = int(emp)
            if e >= 20: score += 2
            elif e >= 5: score += 1
        except: pass
    ca = parse_val(fin.get("current_assets",""))
    if ca and ca > 200_000: score += 1
    return score

def fetch_all_for_sic(sic_code, base_params, api_key):
    auth = "Basic " + b64encode(f"{api_key}:".encode()).decode()
    headers = {"Authorization": auth}
    params_base = {**base_params}
    if sic_code: params_base["sic_codes"] = sic_code
    items = []; start = 0; total = None
    while True:
        params = {**params_base, "size": 100, "start_index": start}
        _rate_limiter.wait_if_needed()
        r = requests.get(API_BASE + "/advanced-search/companies",
                        params=params, headers=headers, timeout=15)
        if r.status_code not in (200,): break
        data = r.json()
        batch = data.get("items", data.get("companies",[]))
        if total is None: total = data.get("hits", data.get("total_results",0))
        if not batch: break
        items.extend(batch)
        start += len(batch)
        if start >= (total or 0) or start >= 5000: break
    return items

def send_email_results(email_to, excel_bytes, csv_data, search_date, criteria):
    import sendgrid
    from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
    sg_key = os.environ.get("SENDGRID_API_KEY","")
    from_email = "sillarsdave@gmail.com"
    body_lines = ["Your Companies House Prospector search has completed.", ""]
    for k, v in criteria.items():
        body_lines.append(f"{k}: {v}")
    body_lines.append("")
    body_lines.append("Please find the Excel and CSV results attached.")
    message = Mail(
        from_email=from_email,
        to_emails=email_to,
        subject=f"Companies House Prospector Results — {search_date}",
        plain_text_content="\n".join(body_lines)
    )
    # Attach Excel
    encoded_xl = base64.b64encode(excel_bytes).decode()
    message.attachment = Attachment(
        FileContent(encoded_xl),
        FileName(f"prospector_results_{search_date}.xlsx"),
        FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        Disposition("attachment")
    )
    # Attach CSV
    csv_bytes = csv_data if isinstance(csv_data, bytes) else csv_data.encode("utf-8-sig")
    encoded_csv = base64.b64encode(csv_bytes).decode()
    message.attachment = Attachment(
        FileContent(encoded_csv),
        FileName(f"prospector_results_{search_date}.csv"),
        FileType("text/csv"),
        Disposition("attachment")
    )
    sg = sendgrid.SendGridAPIClient(api_key=sg_key)
    response = sg.send(message)
    if response.status_code not in (200, 202):
        raise Exception(f"SendGrid error: {response.status_code} {response.body}")

# ─── UI ───────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-header">\U0001f3e2 Companies House Prospector</div>', unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### \U0001f511 API Key")
    try:
        api_key = st.secrets.get("CH_API_KEY") or os.environ.get("CH_API_KEY","")
        if api_key: st.success("API key loaded ✅")
        else:
            api_key = st.text_input("Companies House API Key", type="password",
                                     value=st.session_state.get("api_key",""))
            if api_key: st.session_state["api_key"] = api_key
    except:
        api_key = os.environ.get("CH_API_KEY","")
        if api_key: st.success("API key loaded ✅")
        else:
            api_key = st.text_input("Companies House API Key", type="password",
                                     value=st.session_state.get("api_key",""))
            if api_key: st.session_state["api_key"] = api_key

    st.markdown("---")
    st.markdown("### 📧 Email Results")
    try:
        gmail_user = st.secrets.get("GMAIL_USER") or os.environ.get("GMAIL_USER","")
        gmail_pass = st.secrets.get("GMAIL_APP_PASSWORD") or os.environ.get("GMAIL_APP_PASSWORD","")
        email_configured = bool(gmail_user and gmail_pass)
    except:
        gmail_user = os.environ.get("GMAIL_USER","")
        gmail_pass = os.environ.get("GMAIL_APP_PASSWORD","")
        email_configured = bool(gmail_user and gmail_pass)
    send_email = st.checkbox("Email results when complete", value=email_configured)
    email_to = st.text_input("Send to", value="david.sillars@quilterfa.com")

    st.markdown("---")
    st.markdown("### \U0001f4cd Location")
    location = st.text_input("Location", value="Surrey")

    st.markdown("### \U0001f3ed Industry")
    st.caption("Hold Ctrl/Cmd to select multiple")
    sic_labels = [s[0] for s in SIC_OPTIONS]
    sic_codes  = [s[1] for s in SIC_OPTIONS]
    selected_sic_labels = st.multiselect("Industry / SIC code", sic_labels, default=[])

    st.markdown("### \U0001f3e2 Company Type")
    type_ltd = st.checkbox("Private Limited (Ltd)", value=True)
    type_llp = st.checkbox("LLP", value=True)
    type_plc = st.checkbox("Public (PLC)", value=False)

    st.markdown("### \u2605 Quality Filters")
    excl_dormant = st.checkbox("Exclude dormant", value=True)
    min_age = st.number_input("Min age (years)", value=3, min_value=0, max_value=50)
    max_age = st.number_input("Max age (years)", value=0, min_value=0, max_value=100, help="0 = no limit")
    min_net_assets = st.number_input("Min net assets (\u00a3)", value=0, min_value=0)

    st.markdown("### \U0001f4c8 Financial Data")
    fetch_financials_flag = st.checkbox("Fetch financials (slower)", value=True)
    st.caption("~3s per company. Uncheck for fast search.")

    st.markdown("### \U0001f465 Employee Filter")
    col3, col4 = st.columns(2)
    with col3: emp_min = st.number_input("Min", value=0, min_value=0, key="emp_min")
    with col4: emp_max = st.number_input("Max", value=0, min_value=0, key="emp_max", help="0 = no limit")

    st.markdown("### \u2699\ufe0f Output")
    one_per_company = st.checkbox("One contact per company", value=True)
    linkedin_hyperlinks = st.checkbox("Clickable LinkedIn links", value=True,
                                      help="Uncheck for very large searches (London etc.) to avoid Excel's 65,530 hyperlink limit")

    st.markdown("---")
    search_btn = st.button("\U0001f50d Search Companies House", use_container_width=True)
    if st.button("🗑 Clear Results", use_container_width=True):
        st.session_state.results = []
        st.rerun()

# ─── Main ─────────────────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = []

if search_btn:
    if not api_key:
        st.error("Please enter your API key in the sidebar.")
    elif not selected_sic_labels:
        st.warning("Please select at least one industry.")
    elif not email_to:
        st.warning("Please enter an email address to receive results.")
    else:
        import uuid
        selected_sics = [sic_codes[sic_labels.index(l)] for l in selected_sic_labels]
        selected_types = []
        if type_ltd: selected_types.append("ltd")
        if type_llp: selected_types.append("llp")
        if type_plc: selected_types.append("plc")

        job = {
            "job_id": str(uuid.uuid4()),
            "location": location,
            "sic_codes": selected_sics,
            "sic_labels": selected_sic_labels,
            "company_types": selected_types,
            "fetch_financials": fetch_financials_flag,
            "min_age": int(min_age),
            "max_age": int(max_age),
            "excl_dormant": excl_dormant,
            "min_net_assets": int(min_net_assets),
            "emp_min": int(emp_min),
            "emp_max": int(emp_max),
            "one_per_company": one_per_company,
            "linkedin_hyperlinks": linkedin_hyperlinks,
            "email_to": email_to,
            "submitted_at": time.time(),
        }
        try:
            _r = get_redis()
            # Cancel any currently running job before starting new one
            try:
                _old_status = _r.get("ch_status")
                if _old_status:
                    _old = json.loads(_old_status)
                    if _old.get("running") and _old.get("job_id"):
                        _r.set("ch_cancel", _old.get("job_id"))
            except: pass

            _r.set("ch_job", json.dumps(job))
            _r.set("ch_status", json.dumps({"running": False, "stage": "Queued",
                "job_id": job["job_id"], "email_sent": False, "ready_to_email": False,
                "error": None, "total": 0, "dir_done": 0, "fin_done": 0}))
            _r.delete("ch_results_excel")
            _r.delete("ch_results_csv")
            _r.delete("ch_results_meta")
            _job_ref = job["job_id"][:8].upper()
            st.success(f"✅ Search **{_job_ref}** queued! Results will be emailed to **{email_to}** when complete. You can close this browser.")
            st.info(f"Searching: **{location}** | **{len(selected_sic_labels)}** industries | **{len(selected_sics)}** SIC codes")
            time.sleep(2)
            st.rerun()
        except Exception as e:
            st.error(f"Failed to queue search: {e}")


# ─── Background job status + email trigger ────────────────────────────────────
_status = {}
try:
    _r = get_redis()
    _raw = _r.get("ch_status")
    if _raw:
        _status = json.loads(_raw)
except:
    pass

if _status.get("running"):
    _elapsed = int(time.time() - _status.get("started_at", time.time()))
    _elapsed_str = f"{_elapsed//60}m {_elapsed%60}s" if _elapsed >= 60 else f"{_elapsed}s"
    _total = _status.get("total", 0)
    _d = _status.get("dir_done", 0)
    _fn = _status.get("fin_done", 0)
    _stage = _status.get("stage", "...")

    st.markdown("### 🔍 Search running in background")
    # Show search reference and what search is running
    try:
        _j = json.loads(get_redis().get("ch_job") or "{}")
        _job_ref = _j.get("job_id","")[:8].upper()
        _submitted = _j.get("submitted_at")
        _sub_str = datetime.fromtimestamp(_submitted).strftime("%d %b %Y %H:%M") if _submitted else ""
        st.caption(f"🔖 Search ref: **{_job_ref}** | Submitted: {_sub_str}")
        _loc = _j.get("location", "")
        _inds = ", ".join([l.split("(")[0].strip() for l in _j.get("sic_labels", [])])
        _types = ", ".join([t.upper() for t in _j.get("company_types", [])]) or "All types"
        _min_age = _j.get("min_age", 0)
        _max_age = _j.get("max_age", 0)
        _age_str = f"{_min_age}yr+" if _min_age and not _max_age else (f"{_min_age}-{_max_age}yrs" if _min_age and _max_age else "Any age")
        _mna = _j.get("min_net_assets", 0)
        _mna_str = f"£{_mna:,}+ net assets" if _mna else ""
        _emp_min = _j.get("emp_min", 0)
        _emp_max = _j.get("emp_max", 0)
        _emp_str = f"{_emp_min}-{_emp_max} employees" if _emp_min or _emp_max else ""
        _fin_str = "Financials on" if _j.get("fetch_financials") else "Financials off"
        _dir_str = "1 contact per company" if _j.get("one_per_company") else "All directors"
        _dormant_str = "Excl. dormant" if _j.get("excl_dormant") else "Incl. dormant"

        _line1 = f"📍 {_loc} | 🏭 {_inds}"
        _line2_parts = [_types, _age_str, _dormant_str, _fin_str, _dir_str]
        if _mna_str: _line2_parts.append(_mna_str)
        if _emp_str: _line2_parts.append(_emp_str)
        _line2 = " | ".join(_line2_parts)
        st.caption(_line1)
        st.caption(f"⚙️ {_line2}")
    except: pass
    st.caption(f"Stage: {_stage} | Elapsed: {_elapsed_str}")

    if _total > 0:
        # Directors progress
        _dir_pct = min(_d / _total, 1.0)
        st.markdown(f"**Directors** — {_d:,} of {_total:,}")
        st.progress(_dir_pct)

        # Financials progress
        _fin_pct = min(_fn / _total, 1.0)
        st.markdown(f"**Financials** — {_fn:,} of {_total:,}")
        st.progress(_fin_pct)
    else:
        st.progress(0)
        st.caption("Fetching company list...")

    st.info("✉️ Results will be emailed when complete. You can safely close this browser.")
    st.caption("🔄 Auto-refreshing every 5 seconds...")
    time.sleep(5)
    st.rerun()

elif _status.get("email_sent"):
    # Only show if this status belongs to the current job
    try:
        _current_job = json.loads(get_redis().get("ch_job") or "{}")
        if _status.get("job_id") == _current_job.get("job_id"):
            _job_ref = _current_job.get("job_id","")[:8].upper()
            st.success(f"✅ Search complete — {_status.get('results_count',0):,} results emailed.  (Ref: {_job_ref})")
    except:
        pass

    # Download fallback — show buttons if results are in Redis
    try:
        _r = get_redis()
        _meta_raw = _r.get("ch_results_meta")
        _excel_b64 = _r.get("ch_results_excel")
        _csv_data = _r.get("ch_results_csv")
        if _meta_raw and _excel_b64 and _csv_data:
            _meta = json.loads(_meta_raw)
            _loc = _meta.get("location","").replace(" ","_").lower()[:15]
            _date = _meta.get("search_date","results").replace(" ","_")
            st.markdown("---")
            st.markdown("**📥 Download results directly:**")
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    "⬇️ Download Excel",
                    data=base64.b64decode(_excel_b64),
                    file_name=f"prospector_results_{_loc}_{_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col_dl2:
                st.download_button(
                    "⬇️ Download CSV",
                    data=_csv_data.encode("utf-8-sig"),
                    file_name=f"prospector_results_{_loc}_{_date}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
    except:
        pass

elif _status.get("error"):
    st.error(f"⚠️ Last search failed: {_status.get('error')}")

# ─── Results display ──────────────────────────────────────────────────────────
results = st.session_state.get("results",[])

if results:
    st.markdown(f"### {len(results):,} results loaded")
    rows = []
    for c in results:
        num = c.get("company_number","")
        company_name = title_case_company(c.get("company_name", c.get("title","")))
        addr = c.get("registered_office_address",{})
        addr_str = " ".join(filter(None,[addr.get("address_line_1",""),
                                          addr.get("locality",""), addr.get("postal_code","")]))
        sics = "; ".join(c.get("sic_codes",[]))
        inc = c.get("date_of_creation","")
        age = ""
        if inc:
            try:
                y,m2,d2 = inc.split("-")
                age = str((date.today()-date(int(y),int(m2),int(d2))).days//365)
            except: pass
        fin = st.session_state.get("financials_cache",{}).get(num,{})
        score = calc_score(fin)
        score_str = "\u2605" * min(score,5) if score > 0 else "\u2606"
        dirs = st.session_state.get("director_cache",{}).get(num,[])
        if one_per_company and dirs: dirs = dirs[:1]
        rows_data = dirs if dirs else [None]
        for o in rows_data:
            name = appt = ""
            if o:
                name = " ".join(reversed([p.strip() for p in o.get("name","").split(",")]))
                appt = o.get("appointed_on","")
            first_n, last_n = split_director_name(name)
            ch_url = f"https://find-and-update.company-information.service.gov.uk/company/{num}"
            _sfx_set = {"limited","ltd","llp","plc","and co","company","group","holdings",
                        "holding","services","solutions","consulting","consultancy",
                        "management","associates","partnership","enterprises","ventures",
                        "international","global","uk","the"}
            _co_words = company_name.split() if company_name else []
            if len(_co_words) >= 2 and _co_words[1].lower().rstrip(".") in _sfx_set:
                co_keyword = _co_words[0]
            else:
                co_keyword = " ".join(_co_words[:min(2, len(_co_words))])
            li_url = "https://www.linkedin.com/search/results/people/?keywords=" + requests.utils.quote(f"{first_n} {last_n} {co_keyword}")
            rows.append({
                "Score": score_str, "First Name": first_n, "Surname": last_n,
                "Company": company_name, "Number": num, "Address": addr_str,
                "SIC": sics,
                "Category": ", ".join([l.split("(")[0].strip() for l in selected_sic_labels if any(s in sics for s in [sic_codes[sic_labels.index(l)]])]) or "",
                "Incorporated": inc, "Age": age,
                "Total Assets": fin.get("total_assets",""), "Net Assets": fin.get("net_assets",""),
                "Fixed Assets": fin.get("fixed_assets",""), "Current Assets": fin.get("current_assets",""),
                "Employees": fin.get("employees",""), "Accounts Date": fin.get("accounts_date",""),
                "Dir. Appointed": appt, "Accountant": fin.get("accountant",""),
                "CH Link": ch_url, "LinkedIn": li_url,
            })

    import pandas as pd
    df = pd.DataFrame(rows)
    # Sort by score (desc) then net assets (desc)
    SCORE_ORDER = {"★★★★★": 5, "★★★★": 4, "★★★": 3, "★★": 2, "★": 1, "☆": 0}
    df["_score_num"] = df["Score"].map(SCORE_ORDER).fillna(0)
    df = df.sort_values(["_score_num", "Net Assets"], ascending=[False, False], na_position="last")
    df = df.drop(columns=["_score_num"]).reset_index(drop=True)
    display_df = df.copy()
    display_df["CH company"] = display_df["CH Link"]
    display_df["Officers"] = display_df["CH Link"].apply(lambda x: x + "/officers" if x else "")
    display_df = display_df.drop(columns=["CH Link"])

    st.dataframe(display_df, use_container_width=True, height=600, hide_index=True,
        column_config={
            "Score": st.column_config.TextColumn("★ Score", width=80),
            "First Name": st.column_config.TextColumn("First Name", width=100),
            "Surname": st.column_config.TextColumn("Surname", width=110),
            "Company": st.column_config.TextColumn("Company", width=200),
            "Number": st.column_config.TextColumn("Number", width=90),
            "Address": st.column_config.TextColumn("Address", width=250),
            "SIC": st.column_config.TextColumn("SIC", width=80),
            "Category": st.column_config.TextColumn("Category", width=150),
            "Incorporated": st.column_config.TextColumn("Incorporated", width=100),
            "Age": st.column_config.TextColumn("Age", width=50),
            "Total Assets": st.column_config.TextColumn("Total Assets", width=100),
            "Net Assets": st.column_config.TextColumn("Net Assets", width=100),
            "Fixed Assets": st.column_config.TextColumn("Fixed Assets", width=100),
            "Current Assets": st.column_config.TextColumn("Current Assets", width=110),
            "Employees": st.column_config.TextColumn("Employees", width=85),
            "Accounts Date": st.column_config.TextColumn("Accounts Date", width=110),
            "Dir. Appointed": st.column_config.TextColumn("Dir. Appointed", width=110),
            "Accountant": st.column_config.TextColumn("Accountant", width=180),
            "CH company": st.column_config.LinkColumn("CH Company", width=90, display_text="Open"),
            "Officers": st.column_config.LinkColumn("Officers", width=80, display_text="Officers"),
            "LinkedIn": st.column_config.LinkColumn("LinkedIn", width=80, display_text="Search"),
        })

    def _make_filename(ext):
        loc = location.strip().replace(" ","_").lower()[:15]
        return f"prospector_results_{loc}_{date.today().strftime('%d_%B_%Y')}.{ext}"

    st.markdown("---")
    col_a, col_b, _ = st.columns([1,1,2])
    with col_a:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from collections import Counter
            wb = Workbook(); ws = wb.active; ws.title = "Prospects"
            base_cols = [c for c in df.columns if c not in ["CH Link","LinkedIn"]]
            headers_xl = base_cols + ["CH company","Officers","LinkedIn"]
            hdr_fill = PatternFill("solid", fgColor="1a4a2e")
            for i, h in enumerate(headers_xl, 1):
                cell = ws.cell(row=1, column=i, value=h)
                cell.fill = hdr_fill
                cell.font = Font(name="Arial", color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.row_dimensions[1].height = 30
            for rn, (_, row) in enumerate(df.iterrows(), 2):
                fill = PatternFill("solid", fgColor="EBF3FB") if rn%2==0 else PatternFill("solid", fgColor="FFFFFF")
                row_vals = [row[c] for c in base_cols] + [row["CH Link"], f"{row['CH Link']}/officers", row["LinkedIn"]]
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=rn, column=ci, value=val)
                    cell.fill = fill; cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if ci > len(base_cols):
                        labels = ["Open","Officers","LinkedIn"]
                        cell.value = labels[ci-len(base_cols)-1]
                        cell.hyperlink = val
                        cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            for ci, h in enumerate(headers_xl, 1):
                col_letter = get_column_letter(ci)
                max_len = max(len(str(h)), max((len(str(ws.cell(row=rn,column=ci).value or "")) for rn in range(2,ws.max_row+1)), default=0))
                ws.column_dimensions[col_letter].width = min(max(max_len+2,8),40)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            # Accountants sheet
            from collections import Counter
            ws_acct = wb.create_sheet("Accountants")
            acct_counts = Counter(r for r in df["Accountant"].tolist()
                                  if r and str(r).strip()
                                  and "audit" not in str(r).lower()
                                  and str(r).strip().lower() not in {"n/a", "none", "not applicable"}
                                  and len(str(r).strip()) > 3)
            for ci, h in enumerate(["Accountant Firm", "No. of Clients", "Companies"], 1):
                cell = ws_acct.cell(row=1, column=ci, value=h)
                cell.fill = PatternFill("solid", fgColor="1a4a2e")
                cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            acct_companies = {}
            for _, row in df.iterrows():
                acct = str(row.get("Accountant", "")).strip()
                if acct:
                    acct_companies.setdefault(acct, [])
                    co = str(row.get("Company", "")).strip()
                    if co and co not in acct_companies[acct]:
                        acct_companies[acct].append(co)
            for rn, (acct, count) in enumerate(acct_counts.most_common(), 2):
                ws_acct.cell(row=rn, column=1, value=acct).font = Font(name="Arial", size=9)
                ws_acct.cell(row=rn, column=2, value=count).font = Font(name="Arial", size=9)
                ws_acct.cell(row=rn, column=3, value=", ".join(acct_companies.get(acct, []))).font = Font(name="Arial", size=9)
            ws_acct.column_dimensions["A"].width = 40
            ws_acct.column_dimensions["B"].width = 14
            ws_acct.column_dimensions["C"].width = 60

            # Search Criteria sheet
            ws_crit = wb.create_sheet("Search Criteria")
            _co_types = []
            if type_ltd: _co_types.append("LTD")
            if type_llp: _co_types.append("LLP")
            if type_plc: _co_types.append("PLC")
            _age_str = f"{int(min_age)}yr+" if min_age and not max_age else (f"{int(min_age)}-{int(max_age)}yrs" if min_age and max_age else "Any")
            _emp_str = f"{int(emp_min)}-{int(emp_max)}" if (emp_min or emp_max) else "Any"
            criteria_data = {
                "Location": location,
                "Industries": ", ".join(selected_sic_labels),
                "Company types": ", ".join(_co_types) if _co_types else "All",
                "Min age": _age_str,
                "Exclude dormant": "Yes" if excl_dormant else "No",
                "Min net assets": f"£{int(min_net_assets):,}" if min_net_assets else "None",
                "Employees": _emp_str,
                "Fetch financials": "Yes" if fetch_financials_flag else "No",
                "One contact per company": "Yes" if one_per_company else "No",
                "Results in export": f"{len(rows):,}",
                "Export date": date.today().strftime("%d %B %Y"),
            }
            for i, (k, v) in enumerate(criteria_data.items(), 1):
                ws_crit.cell(row=i, column=1, value=k).font = Font(bold=True, name="Arial")
                ws_crit.cell(row=i, column=2, value=str(v)).font = Font(name="Arial")
            ws_crit.column_dimensions["A"].width = 28
            ws_crit.column_dimensions["B"].width = 50
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("\u2b07 Download Excel", data=buf.getvalue(),
                               file_name=_make_filename("xlsx"),
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        except Exception as e:
            st.error(f"Excel error: {e}")

    with col_b:
        df_csv = df.copy()
        df_csv["CH company"] = df["CH Link"]
        df_csv["Officers"] = df["CH Link"].apply(lambda x: x+"/officers")
        df_csv["LinkedIn search"] = df["LinkedIn"]
        df_csv = df_csv.drop(columns=["CH Link","LinkedIn"])
        st.download_button("\u2b07 Download CSV", data=df_csv.to_csv(index=False).encode("utf-8-sig"),
                           file_name=_make_filename("csv"), mime="text/csv", use_container_width=True)

else:
    st.info("Configure your filters in the sidebar and click **Search Companies House** to begin.")

# ─── Always auto-refresh if job running or ready to email ─────────────────────
try:
    _ar = get_redis()
    _ar_raw = _ar.get("ch_status")
    if _ar_raw:
        _ar_status = json.loads(_ar_raw)
        if _ar_status.get("running") or (_ar_status.get("ready_to_email") and not _ar_status.get("email_sent")):
            time.sleep(5)
            st.rerun()
except:
    pass
